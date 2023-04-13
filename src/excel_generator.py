from pathlib import Path
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Side, Border, NamedStyle, Font, colors
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter

from data_processor import DataProcessor

centered_alignment = Alignment(horizontal='center', vertical='center')
thin_side = Side(border_style="thin", color="000000")
double_side = Side(border_style="double", color="000000")
thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
bold_font = Font(bold=True)
centered_thin_border_style = NamedStyle(name='CenteredThinBorder', alignment=centered_alignment, border=thin_border)
centered_bold_thin_border_style = NamedStyle(name='CenteredBoldThinBorder', font=bold_font,
                                             alignment=centered_alignment, border=thin_border)
bold_thin_border_style = NamedStyle(name='BoldThinBorder', font=bold_font, border=thin_border)
thin_border_style = NamedStyle(name='ThinBorder', border=thin_border)
thin_border_pct_style = NamedStyle(name='ThinBorderPct', border=thin_border, number_format=FORMAT_PERCENTAGE_00)

centered_bold_style = NamedStyle(name='CenteredBold', alignment=centered_alignment, font=bold_font)
ryg_color_scale_rule = ColorScaleRule(start_type='percentile', start_value=2.5, start_color='f44336',
                                      mid_type='percentile', mid_value=50, mid_color='fff9c4',
                                      end_type='percentile', end_value=97.5, end_color='00a933')
rg_color_scale_rule = ColorScaleRule(start_type='percentile', start_value=2.5, start_color='f44336',
                                     end_type='percentile', end_value=97.5, end_color='00a933')
ry_color_scale_rule = ColorScaleRule(start_type='percentile', start_value=2.5, start_color='f44336',
                                     end_type='percentile', end_value=97.5, end_color='00a933')
gy_color_scale_rule = ColorScaleRule(start_type='percentile', start_value=2.5, start_color='fff9c4',
                                     end_type='percentile', end_value=97.5, end_color='f44336')
blue_bar_rule = DataBarRule(start_type='min', end_type='max', color=colors.BLUE)


def set_rows_height(ws, row_height, table_first_row, table_last_row):
    """
    Sets height for the rows in a range.

    :param ws: active worksheet.
    :param row_height: new row's height.
    :param table_first_row: [int] first row from the range.
    :param table_last_row: [int] last row from the range.
    :return: None
    """
    for row_idx in range(table_first_row, table_last_row + 1):
        ws.row_dimensions[row_idx].height = row_height


def set_columns_width(ws, column_width, first_column_idx, last_column_idx):
    """
    Sets width for the columns in a range.

    :param ws: active worksheet.
    :param column_width: new column's width.
    :param first_column_idx: [int] first column of the range.
    :param last_column_idx: [int] first column of the range.
    :return: None
    """
    for col_idx in range(first_column_idx, last_column_idx + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = column_width


def set_style(ws, cell_range, style: NamedStyle):
    """
    Sets a style to all the cell in range.
    The range is declared using worksheet coordinates, e.g. "A1:J10"

    :param ws: active worksheet.
    :param cell_range: cell range to apply the style.
    :param style: a NamedStyle object.
    :return: None
    """
    for row in ws[cell_range]:
        for cell in row:
            cell.style = style


def insert_table_tittle(ws, tittle, tittle_width=1):
    """
    Insert a tittle with headline style, merge cells and centers the text.
    This is useful to center the title above a table.

    :param ws: active worksheet.
    :param tittle: the tittle.
    :param tittle_width: how many cells will be merged with the tittle
    :return: None
    """
    ws.append([tittle])
    tittle_row = ws.max_row
    ws[f'A{tittle_row}'].style = 'Headline 1'
    table_last_column = get_column_letter(tittle_width)
    ws.merge_cells(f'A{tittle_row}:{table_last_column}{tittle_row}')
    ws[f'A{tittle_row}'].alignment = centered_alignment
    ws.row_dimensions[tittle_row].height = 25


def create_big_matrix_sheet(wb, sheet_name, table: pd.DataFrame, color_scale_rule):
    """
    Creates a worksheet and inserts a matrix with a conditional formatting rule.

    :param wb: active workbook.
    :param sheet_name: the new sheet's name.
    :param table: a pandas Dataframe with the data to insert.
    :param color_scale_rule: a color scale rule to apply to the table.
    :return: None
    """
    ws = wb.create_sheet(title=sheet_name)
    insert_formatted_matrix(ws, table, color_scale_rule)


def insert_formatted_matrix(ws, table: pd.DataFrame, color_scale_rule):
    """
    Inserts a table in the given worksheet and applies a conditional formatting rule.

    :param ws: active worksheet.
    :param table: a pandas Dataframe with the data to insert.
    :param color_scale_rule: a color scale rule to apply to the table.
    :return: None
    """
    ws.append(['', *table.columns.to_list()])
    table_first_row = ws.max_row
    for row in table.iterrows():
        ws.append([row[0], *row[1].values])
    table_headers_range = f'A{table_first_row}:{get_column_letter(ws.max_column)}{table_first_row}'
    set_style(ws, table_headers_range, centered_bold_style)
    table_index_range = f'A{table_first_row}:A{ws.max_row}'
    set_style(ws, table_index_range, centered_bold_style)
    set_columns_width(ws, 15, 2, ws.max_column)
    ws.conditional_formatting.add(f'B{table_first_row + 1}:{get_column_letter(ws.max_column)}{ws.max_row}',
                                  color_scale_rule)


class ExcelGenerator:
    """
    From data processed by the DataProcessor, it creates several excel files in a give destination folder.

    There are two kinds of excel files:
        - 'all_futures_tables.xls': this excel contains aggregated data of all futures processed by the DataProcessor
          instance. This file contains the following sheets:
             * 1_AllFutures: a table with all the futures to be used as index.
             * 2_MovementByHour: a table with the  normalized mena movement by hour for all futures.
             * 3_AbsoluteMovementByHour: similar to the previous table, but this one contains the absolute movement.
               It's a proxy of the movement strength by hour.
             * 4_NormalizedMovementByHour: this is plot with the mean of all columns in the previous table,
               normalized to 1. This plot summarizes the movement strength across all data processed by the DataProcessor.
             * 5_CorrelationMatrix: the correlation matrix of all futures in a huge table.
        - Future-specific workbook: for all futures in the processed data, we create a workbook, it contains:
            * Correlation matrix of the top 10 strongest correlated futures with the selected future.
            * Correlation matrix of the top 10 weakest correlated futures with the selected future.
            * Statistics of all positive and negatives days in the complete data set.
            * Mean movement in USDT by hours and its strength.
    """
    def __init__(self, destination_folder: Path, data_processor: DataProcessor):
        """
        Initializes an instance of the ExcelGenerator class.

        :param destination_folder: the folder where all excel files will be saved. If the folder doesn't exist,
                                   it is created.
        :param data_processor: an instance of DataProcessor.
        """
        if not destination_folder.exists():
            destination_folder.mkdir(parents=True, exist_ok=True)
        print(f'ExcelGenerator: initializing, output data will be saved in {destination_folder.absolute()}.')
        self.destination_folder = destination_folder
        self.data_processor = data_processor
        self.positive_negative_days_statistics = data_processor.estimate_positive_negative_days_statistics()
        self.mean_movement_and_strength_by_hour = data_processor.estimate_mean_movement_and_strength_by_hour()

    def run(self):
        """
        Generates all excel files and saves them in the destination folder.

        :return: None
        """
        self._generate_all_futures_tables_workbook()
        for target in self.data_processor.futures_list:
            self._generate_future_specific_workbook(target)

    def _generate_all_futures_tables_workbook(self):
        """
        Generates the 'all_futures_tables.xls' workbook and saves it in the destination folder.

        :return: None
        """
        wb = Workbook()
        # remove default sheet
        wb.remove(wb.active)
        self._create_futures_index_sheet(wb)
        self._create_movement_by_hour_sheet(wb)
        self._create_absolute_movement_by_hour_sheet(wb)
        self._create_normalized_movement_by_hour_sheet(wb)
        self._create_correlation_matrix_sheet(wb)
        self._create_unstacked_correlation_matrix_sheet(wb)
        xlsx_file_path = self.destination_folder / "all_futures_tables.xlsx"
        wb.save(xlsx_file_path.resolve())
        print(f'ExcelGenerator: saving all futures workbook at {xlsx_file_path.absolute()}.')
        wb.close()

    def _generate_future_specific_workbook(self, target):
        """
        Generates the future-specific workbook for the target future and saves it in the destination folder.

        :param target: the specific future.
        :return: None
        """
        wb = Workbook()
        # remove default sheet
        wb.remove(wb.active)
        ws = wb.create_sheet(title=target)
        top_count = 10
        self._insert_correlation_matrices(ws, target, top_count)
        self._insert_positive_negative_statistics(ws, target)
        self._insert_movement_by_hour(ws, target)
        xlsx_file_path = self.destination_folder / f'{target}.xlsx'
        wb.save(xlsx_file_path.resolve())
        print(f'ExcelGenerator: saving workbook for {target} at {xlsx_file_path.absolute()}.')
        wb.close()

    def _insert_movement_by_hour(self, ws, target):
        """
        Inserts the mean movement in USDT by hours and its strength for the target future in the future-specific workbook.

        :param ws: active worksheet.
        :param target: the specific future.
        :return: None
        """
        insert_table_tittle(ws, 'Mean Movement by Hour', 3)
        table = self.mean_movement_and_strength_by_hour[target]
        ws.append(['Hour', *table.columns.to_list()])
        table_first_row = ws.max_row
        for row in table.iterrows():
            ws.append([row[0], *row[1].values])
        table_last_row = ws.max_row
        # Format table
        table_headers_range = f'A{table_first_row}:C{table_first_row}'
        set_style(ws, table_headers_range, centered_bold_style)
        table_index_range = f'A{table_first_row}:A{table_last_row}'
        set_style(ws, table_index_range, centered_bold_style)
        mean_movement_range = f'B{table_first_row + 1}:B{table_last_row}'
        ws.conditional_formatting.add(mean_movement_range, ryg_color_scale_rule)
        movement_strength_range = f'C{table_first_row + 1}:C{table_last_row}'
        ws.conditional_formatting.add(movement_strength_range, blue_bar_rule)
        set_columns_width(ws, 25, 2, 3)

    def _insert_positive_negative_statistics(self, ws, target):
        """
        Inserts the statistics of all positive and negatives days for the target future in the future-specific workbook.

        :param ws: Active worksheet.
        :param target: the specific future.
        :return: None
        """
        # Insert Positive and Negative days statistics
        insert_table_tittle(ws, 'Positive and Negative days statistics', 3)
        table = self.positive_negative_days_statistics[target]
        ws.append(['', *table.columns.to_list()])
        table_first_row = ws.max_row
        for row in table.iterrows():
            ws.append([row[0], *row[1].values])
        table_last_row = ws.max_row
        set_columns_width(ws, 27, 1, 1)
        # Format table
        table_range = f'A{table_first_row}:C{table_last_row}'
        set_style(ws, table_range, thin_border_style)
        table_headers_range = f'A{table_first_row}:C{table_first_row}'
        set_style(ws, table_headers_range, centered_bold_thin_border_style)
        table_index_range = f'A{table_first_row}:A{table_last_row}'
        set_style(ws, table_index_range, bold_thin_border_style)
        pct_range = f'B{table_first_row + 2}:C{table_last_row - 1}'
        set_style(ws, pct_range, thin_border_pct_style)
        ws.append([''])

    def _insert_correlation_matrices(self, ws, target, top_count):
        """
        Inserts the correlation matrices for the top strongest and top weakest correlated futures against the target future.

        :param ws: active worksheet.
        :param target: the specific future.
        :param top_count: the count of the top strongest (weakest) correlated futures.
        :return: None
        """
        # The top 10  + index column + the target future
        matrices_cell_width = top_count + 2
        insert_table_tittle(ws, target, matrices_cell_width)
        ws.append([''])
        # Insert correlation matrices
        corr_matrices = self.data_processor.get_correlation_matrices_respect_to(target, top_count)
        insert_table_tittle(ws, f'Strongest correlations with {target}', matrices_cell_width)
        insert_formatted_matrix(ws, corr_matrices['HighestCorrelated'], ry_color_scale_rule)
        ws.append([''])
        insert_table_tittle(ws, f'Weakest correlations with {target}', matrices_cell_width)
        insert_formatted_matrix(ws, corr_matrices['LowestCorrelated'], gy_color_scale_rule)
        ws.append([''])

    def _create_futures_index_sheet(self, wb, batch_size=10):
        """
        Creates the 1_AllFutures in the 'all_futures_tables.xls' file.

        :param wb: active workbook.
        :param batch_size: the number of columns each row contains. Default: 10.
        :return: None
        """
        ws = wb.create_sheet(title='1_AllFutures')
        insert_table_tittle(ws, 'FUTUROS BINANCE', batch_size)
        # Generate index table.
        table_first_row = ws.max_row + 1
        futures_list = self.data_processor.futures_list
        for row_idx, idx in enumerate(range(0, len(futures_list), batch_size), start=table_first_row):
            for column_idx, pair_idx in enumerate(range(idx, idx + batch_size), start=1):
                pair = futures_list[pair_idx]
                coords = f'{get_column_letter(column_idx)}{row_idx}'
                ws[coords] = pair
                ws[coords].alignment = centered_alignment
                if pair_idx == len(futures_list) - 1:
                    table_last_row = row_idx
                    break
        # Format table
        set_columns_width(ws, 15, 1, ws.max_column)
        set_rows_height(ws, 20, table_first_row, table_last_row)
        cell_range = f'A{table_first_row}:{get_column_letter(ws.max_column)}{table_last_row}'
        set_style(ws, cell_range, centered_thin_border_style)

    def _create_movement_by_hour_sheet(self, wb):
        """
        Creates the 2_MovementByHour in the 'all_futures_tables.xls' file.

        :param wb: active workbook.
        :return: None
        """
        table = self.data_processor.estimate_normalized_mean_movement_by_hour()
        create_big_matrix_sheet(wb, '2_MovementByHour', table, ryg_color_scale_rule)

    def _create_absolute_movement_by_hour_sheet(self, wb):
        """
        Creates the 3_AbsoluteMovementByHour in the 'all_futures_tables.xls' file.

        :param wb: active workbook.
        :return: None
        """
        table = self.data_processor.estimate_normalized_absolute_mean_movement_by_hour()
        create_big_matrix_sheet(wb, '3_AbsoluteMovementByHour', table, rg_color_scale_rule)

    def _create_normalized_movement_by_hour_sheet(self, wb):
        """
        Creates the 4_NormalizedMovementByHour in the 'all_futures_tables.xls' file.

        :param wb: active workbook.
        :return: None
        """
        normalized_movement_hour = self.data_processor.estimate_normalized_absolute_mean_movement_by_hour()
        mean_absolute_movement_by_hour = (normalized_movement_hour.transpose().sum().mean())
        ax = ((normalized_movement_hour.transpose().sum() / mean_absolute_movement_by_hour)
              .plot(kind='barh', figsize=(20, 15), grid=True, fontsize=18))
        title = 'Normalized movement strength by hours'
        ax.set_title(title, pad=20, fontdict={'fontsize': 24})
        ax.set_ylabel('HOURS', fontdict={'fontsize': 20})
        fig_name = NamedTemporaryFile()
        ax.figure.tight_layout()
        ax.figure.savefig(fig_name)
        ws = wb.create_sheet(title="4_NormalizedMovementByHour")
        img = Image(fig_name)
        img.anchor = 'A1'
        ws.add_image(img)

    def _create_correlation_matrix_sheet(self, wb):
        """
        Creates the 5_CorrelationMatrix in the 'all_futures_tables.xls' file.

        :param wb: active workbook.
        :return: None
        """
        table = self.data_processor.correlation_matrix
        sheet_name = '5_CorrelationMatrix'
        create_big_matrix_sheet(wb, sheet_name, table, rg_color_scale_rule)
        set_columns_width(wb[sheet_name], 15, 1, 1)

    def _create_unstacked_correlation_matrix_sheet(self, wb, hidden=True):
        """
        Creates the UnstackedCorrelationMatrix in the 'all_futures_tables.xls' file. By default, this sheet is hidden
        because it is used as source of the DynamicCorrelationMatrix pivot table.

        :param wb: active workbook.
        :param hidden: if True, the sheet will be hidden. Default: True
        :return: None
        """
        ws = wb.create_sheet(title="99_UnstackedCorrelationMatrix")
        ws.append(['THIS', 'OTHER', 'Correlation'])
        unstacked_correlation = self.data_processor.correlation_matrix.unstack()
        for row in unstacked_correlation.items():
            ws.append([*row[0], row[1]])
        if hidden:
            ws.sheet_state = 'hidden'
